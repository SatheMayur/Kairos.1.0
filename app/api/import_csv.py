"""CSV / batch import endpoint — feed Naukri & WorkIndia downloads into the pipeline.

POST /api/v1/import/csv
  - Upload a CSV file from Naukri or WorkIndia employer dashboard
  - System parses, deduplicates, scores, shortlists, and queues outreach automatically

POST /api/v1/import/batch
  - Submit a JSON array of candidate objects (for manual / scripted entry)

Both endpoints return a per-candidate summary: scored, shortlisted, outreach queued.
"""
import json
import re
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.adapters.base import RawCandidate
from app.adapters.naukri import NaukriCSVAdapter
from app.adapters.workindia import WorkIndiaCSVAdapter
from app.adapters.apna import ApnaCSVAdapter
from app.api.deps import get_db
from app.config import get_settings
from app.models.candidate import CandidateSource
from app.models.outreach import OutreachChannel, OutreachType
from app.models.shortlist import ShortlistStatus
from app.services.outreach import send_outreach
from app.services.sourcing import _score_and_shortlist, _upsert_candidate
from app.utils.logging import get_logger

logger = get_logger(__name__)
settings = get_settings()
router = APIRouter(prefix="/import", tags=["import"])

_NAUKRI_ADAPTER = NaukriCSVAdapter()
_WORKINDIA_ADAPTER = WorkIndiaCSVAdapter()
_APNA_ADAPTER = ApnaCSVAdapter()


def _xlsx_to_csv(content: bytes) -> str:
    """Convert an uploaded Excel (.xlsx) file into CSV text so the existing
    CSV adapters can parse it. Reads the first/active sheet.

    Some exporters (e.g. Apna) write a wrong/missing sheet 'dimension', which
    makes openpyxl's fast read-only mode stop after one row. So we read in
    read-only mode first and, if that yields no data rows, fall back to the
    full (non-read-only) reader which scans every row regardless of dimension.
    """
    import csv as _csv
    import io as _io
    from openpyxl import load_workbook

    def _read(read_only: bool) -> list[list[str]]:
        wb = load_workbook(_io.BytesIO(content), read_only=read_only, data_only=True)
        ws = wb.active
        rows = [["" if c is None else str(c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows

    rows = _read(True)
    if len(rows) < 2:                 # truncated by a bad dimension tag → full read
        rows = _read(False)

    out = _io.StringIO()
    writer = _csv.writer(out)
    for row in rows:
        writer.writerow(row)
    return out.getvalue()


class BatchCandidate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    skills: list[str] = []
    experience_years: Optional[float] = None
    current_salary: Optional[float] = None
    expected_salary: Optional[float] = None
    location: Optional[str] = None
    notice_period_days: Optional[int] = None
    current_employer: Optional[str] = None
    current_role: Optional[str] = None
    source: str = "MANUAL"


class ImportResult(BaseModel):
    total_parsed: int
    inserted: int
    duplicates_skipped: int
    auto_shortlisted: int
    pending_review: int
    rejected: int
    outreach_queued: int
    details: list[dict]


async def _run_import_pipeline(
    raw_candidates: list[RawCandidate],
    job_id: int,
    auto_outreach: bool,
    db: AsyncSession,
) -> ImportResult:
    """Score every candidate, shortlist, and optionally queue outreach."""
    from sqlalchemy import select
    from app.models.job import Job

    job_res = await db.execute(select(Job).where(Job.id == job_id))
    job = job_res.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    inserted = 0
    duplicates = 0
    auto_sl = 0
    pending = 0
    rejected = 0
    outreach_queued = 0
    details = []

    for raw in raw_candidates:
        existing_before = True
        try:
            from sqlalchemy import select as sa_select
            from app.models.candidate import Candidate
            exists = False
            if raw.email:
                r = await db.execute(
                    sa_select(Candidate).where(Candidate.email == raw.email)
                )
                exists = r.scalar_one_or_none() is not None
            if not exists and raw.source_ref:
                r = await db.execute(
                    sa_select(Candidate).where(Candidate.source_ref == raw.source_ref)
                )
                exists = r.scalar_one_or_none() is not None
            if exists:
                duplicates += 1
                details.append({"name": raw.name, "result": "DUPLICATE_SKIPPED"})
                continue
            existing_before = False
        except Exception:
            pass

        candidate = await _upsert_candidate(raw, db)
        if not existing_before:
            inserted += 1

        entry = await _score_and_shortlist(candidate, job, db)
        if not entry:
            details.append({"name": raw.name, "result": "ALREADY_SCORED"})
            continue

        if entry.status == ShortlistStatus.SHORTLISTED:
            auto_sl += 1
            result_label = "AUTO_SHORTLISTED"
        elif entry.status == ShortlistStatus.PENDING:
            pending += 1
            result_label = "PENDING_REVIEW"
        else:
            rejected += 1
            result_label = "REJECTED"

        # Auto-WhatsApp every reachable candidate the moment they're imported —
        # shortlisted AND pending-review — as long as we can actually reach them
        # (a valid mobile or a real email). Phone-less ones (e.g. locked Apna
        # profiles) are skipped here and shown in Needs Fixing to unlock; you
        # can't message a hidden number.
        from app.services.data_quality import is_reachable
        if (auto_outreach
                and entry.status in (ShortlistStatus.SHORTLISTED, ShortlistStatus.PENDING)
                and is_reachable(candidate)):
            from app.models.shortlist import ShortlistEntry
            try:
                log = await send_outreach(
                    candidate=candidate,
                    job=job,
                    channel=OutreachChannel.WHATSAPP,
                    outreach_type=OutreachType.INITIAL_CONTACT,
                    db=db,
                )
                # Only count as contacted when a REAL channel delivered AND the
                # candidate is genuinely reachable. A PLATFORM_MESSAGE/UNREACHABLE
                # "send" (e.g. a phone-locked Apna candidate whose only address is
                # apna:<id>) logs SENT but reaches no one — marking it CONTACTED
                # would strand the candidate.
                if (log.status.value == "SENT"
                        and log.channel in (
                            OutreachChannel.WHATSAPP, OutreachChannel.EMAIL, OutreachChannel.SMS
                        )
                        and is_reachable(candidate)):
                    outreach_queued += 1
                    entry.status = ShortlistStatus.CONTACTED
            except Exception as exc:
                logger.warning("Outreach failed for %s: %s", candidate.name, exc)

        breakdown = entry.score_breakdown or {}
        details.append({
            "name": raw.name,
            "email": raw.email,
            "score": entry.score,
            "result": result_label,
            "candidate_id": candidate.id,
            "strengths": breakdown.get("ai_strengths", []),
            "concerns": breakdown.get("ai_concerns", []),
            "reasoning": breakdown.get("ai_reasoning", ""),
            "opener": breakdown.get("ai_opener", ""),
        })

    await db.commit()

    return ImportResult(
        total_parsed=len(raw_candidates),
        inserted=inserted,
        duplicates_skipped=duplicates,
        auto_shortlisted=auto_sl,
        pending_review=pending,
        rejected=rejected,
        outreach_queued=outreach_queued,
        details=details,
    )


@router.post("/csv", response_model=ImportResult)
async def import_csv(
    file: UploadFile = File(..., description="CSV or Excel (.xlsx) export from Naukri, Apna, or WorkIndia"),
    job_id: int = Form(..., description="Job ID to score candidates against"),
    source: str = Form("NAUKRI", description="Portal source: NAUKRI or WORKINDIA"),
    auto_outreach: bool = Form(True, description="Queue outreach for AUTO_SHORTLIST candidates"),
    db: AsyncSession = Depends(get_db),
):
    """Upload a Naukri or WorkIndia CSV export."""
    src = source.upper().strip()
    if src not in ("NAUKRI", "WORKINDIA", "APNA"):
        raise HTTPException(status_code=400, detail="source must be NAUKRI, WORKINDIA or APNA")

    content = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith(".xlsx") or content[:2] == b"PK":
        try:
            csv_text = _xlsx_to_csv(content)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=("Could not read this Excel file. If it's an old .xls, open it and "
                        f"'Save As' .xlsx or .csv, then upload again. ({str(exc)[:100]})"),
            )
    else:
        try:
            csv_text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            csv_text = content.decode("latin-1")

    if src == "NAUKRI":
        raw_candidates = _NAUKRI_ADAPTER.parse_csv(csv_text)
    elif src == "APNA":
        raw_candidates = _APNA_ADAPTER.parse_csv(csv_text)
    else:
        raw_candidates = _WORKINDIA_ADAPTER.parse_csv(csv_text)

    if not raw_candidates:
        raise HTTPException(status_code=400, detail="No candidates found in the file — check the format")

    logger.info("CSV import: source=%s job=%d candidates=%d", src, job_id, len(raw_candidates))
    return await _run_import_pipeline(raw_candidates, job_id, auto_outreach, db)


@router.post("/apna", response_model=ImportResult)
async def import_apna(
    file: UploadFile = File(..., description="Apna 'Download Excel' export"),
    job_title: str = Form(..., description="The Apna job's title — matched to (or used to create) a role here"),
    company: str = Form("Bookends Hospitality", description="Company the Apna account hires for"),
    auto_outreach: bool = Form(True),
    db: AsyncSession = Depends(get_db),
):
    """Used by the local Apna Sync helper."""
    from sqlalchemy import select, func
    from app.models.job import Job, JobStatus

    title = (job_title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="job_title is required")

    # Match on normalized title (collapse inner whitespace, case-insensitive) AND
    # company, so "Sr. Graphic Designer" doesn't attach to another company's role
    # and "Sr Graphic Designer" / double-spaced variants don't spawn duplicate jobs.
    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip()).lower()

    norm_title, norm_company = _norm(title), _norm(company)
    all_jobs = (await db.execute(select(Job))).scalars().all()
    job = next(
        (j for j in all_jobs
         if _norm(j.title) == norm_title and (not norm_company or _norm(j.company) == norm_company)),
        None,
    )
    if not job:
        job = Job(title=title, company=company or None, status=JobStatus.ACTIVE)
        db.add(job)
        await db.commit()
        await db.refresh(job)
        logger.info("import_apna: auto-created job '%s' (%s) (#%d)", title, company, job.id)

    content = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith(".xlsx") or content[:2] == b"PK":
        try:
            csv_text = _xlsx_to_csv(content)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not read this Excel file. ({str(exc)[:100]})")
    else:
        try:
            csv_text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            csv_text = content.decode("latin-1")

    raw_candidates = _APNA_ADAPTER.parse_csv(csv_text)
    if not raw_candidates:
        raise HTTPException(status_code=400, detail="No candidates found in the file — check the format")

    logger.info("Apna import: job='%s'(#%d) candidates=%d", title, job.id, len(raw_candidates))
    return await _run_import_pipeline(raw_candidates, job.id, auto_outreach, db)


@router.post("/batch", response_model=ImportResult)
async def import_batch(
    candidates: list[BatchCandidate],
    job_id: int,
    auto_outreach: bool = True,
    db: AsyncSession = Depends(get_db),
):
    """Submit candidates as a JSON array."""
    raw: list[RawCandidate] = []
    for c in candidates:
        try:
            src = CandidateSource(c.source.upper())
        except ValueError:
            src = CandidateSource.MANUAL
        raw.append(
            RawCandidate(
                name=c.name,
                source=src,
                email=c.email,
                phone=c.phone,
                skills=c.skills,
                experience_years=c.experience_years,
                current_salary=c.current_salary,
                expected_salary=c.expected_salary,
                location=c.location,
                notice_period_days=c.notice_period_days,
                current_employer=c.current_employer,
                current_role=c.current_role,
                source_ref=f"{src.value.lower()}:{c.email or c.phone or c.name}",
            )
        )

    logger.info("Batch import: job=%d candidates=%d", job_id, len(raw))
    return await _run_import_pipeline(raw, job_id, auto_outreach, db)


# ── Smart URL import (ScrapeGraph-style) ──────────────────────────────────────

class URLImportRequest(BaseModel):
    url: str
    job_id: int
    source: str = "MANUAL"
    auto_outreach: bool = True


class BulkURLImportRequest(BaseModel):
    urls: list[str]
    job_id: int
    source: str = "MANUAL"
    auto_outreach: bool = True


@router.post("/url")
async def import_from_url(payload: URLImportRequest, db: AsyncSession = Depends(get_db)):
    """Scrape any candidate profile URL and import into the pipeline."""
    from app.adapters.smart_scrape import fetch_and_extract
    from app.models.candidate import CandidateSource as CS

    try:
        src = CS(payload.source.upper())
    except ValueError:
        src = CS.MANUAL

    raw = await fetch_and_extract(payload.url, source=src)
    if not raw:
        raise HTTPException(
            status_code=422,
            detail="Could not extract candidate data from that URL. "
                   "Check the URL is a public profile page and ANTHROPIC_API_KEY is set."
        )

    result = await _run_import_pipeline([raw], payload.job_id, payload.auto_outreach, db)
    return {
        "url": payload.url,
        "name": raw.name,
        "skills_found": len(raw.skills or []),
        **result.model_dump(),
    }


@router.post("/url/bulk")
async def import_from_urls(payload: BulkURLImportRequest, db: AsyncSession = Depends(get_db)):
    """Scrape multiple profile URLs concurrently and import all into the pipeline."""
    import asyncio
    from app.adapters.smart_scrape import fetch_and_extract
    from app.models.candidate import CandidateSource as CS

    if len(payload.urls) > 20:
        raise HTTPException(status_code=422, detail="Maximum 20 URLs per request")

    try:
        src = CS(payload.source.upper())
    except ValueError:
        src = CS.MANUAL

    tasks = [fetch_and_extract(url, source=src) for url in payload.urls]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    raw_candidates = []
    failed = []
    for url, res in zip(payload.urls, results):
        if isinstance(res, Exception) or res is None:
            failed.append(url)
        else:
            raw_candidates.append(res)

    if not raw_candidates:
        raise HTTPException(
            status_code=422,
            detail=f"Could not extract data from any of the {len(payload.urls)} URLs."
        )

    import_result = await _run_import_pipeline(
        raw_candidates, payload.job_id, payload.auto_outreach, db
    )
    return {
        "urls_attempted": len(payload.urls),
        "urls_extracted": len(raw_candidates),
        "urls_failed": failed,
        **import_result.model_dump(),
    }


@router.post("/enrich")
async def enrich_existing_candidates(
    job_id: int,
    limit: int = 20,
    db: AsyncSession = Depends(get_db),
):
    """Re-scrape source_ref URLs for existing candidates to fill in missing skills/data."""
    import asyncio
    from sqlalchemy import select
    from app.adapters.smart_scrape import fetch_and_extract
    from app.models.candidate import Candidate
    from app.models.shortlist import ShortlistEntry

    sl_res = await db.execute(
        select(ShortlistEntry).where(ShortlistEntry.job_id == job_id)
    )
    entries = sl_res.scalars().all()
    cand_ids = [e.candidate_id for e in entries]

    if not cand_ids:
        return {"enriched": 0, "skipped": 0, "message": "No candidates for this job"}

    c_res = await db.execute(
        select(Candidate).where(Candidate.id.in_(cand_ids))
    )
    candidates = c_res.scalars().all()

    to_enrich = [
        c for c in candidates
        if c.source_ref
        and c.source_ref.startswith("http")
        and len(c.skills or []) < 3
    ][:limit]

    if not to_enrich:
        return {"enriched": 0, "skipped": len(candidates),
                "message": "No candidates need enrichment (all have skills or no URL)"}

    enriched = skipped = 0
    tasks = [fetch_and_extract(c.source_ref, source=c.source) for c in to_enrich]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    for candidate, raw in zip(to_enrich, results):
        if isinstance(raw, Exception) or raw is None:
            skipped += 1
            continue
        if raw.skills:
            candidate.skills = raw.skills
        if raw.experience_years and not candidate.experience_years:
            candidate.experience_years = raw.experience_years
        if raw.education and not candidate.education:
            candidate.education = raw.education
        if raw.current_role and not candidate.current_role:
            candidate.current_role = raw.current_role
        if raw.phone and not candidate.phone:
            candidate.phone = raw.phone
            candidate.whatsapp = raw.phone
        enriched += 1
        logger.info("Enriched candidate %d (%s) from %s", candidate.id, candidate.name, candidate.source_ref)

    return {"enriched": enriched, "skipped": skipped, "total_candidates": len(candidates)}




# ── Apna Hire live search endpoints ──────────────────────────────────────────

class ApnaSearchRequest(BaseModel):
    keywords: list[str]
    location: Optional[str] = None
    experience_min: Optional[float] = None
    experience_max: Optional[float] = None
    page: int = 1
    size: int = 20
    apna_token: str
    org_id: str = "2012727"


class ApnaImportRequest(BaseModel):
    candidates: list[dict]
    job_id: int
    auto_outreach: bool = True


class ApnaTokenRequest(BaseModel):
    token: str
    org_id: Optional[str] = None


@router.post("/apna-token")
async def save_apna_token(payload: ApnaTokenRequest, db: AsyncSession = Depends(get_db)):
    """Save the Apna sign-in token on the server so automatic, behind-the-scenes
    candidate finding can use Apna too (not just the manual search box).

    The token itself is never sent back or logged.
    """
    from app.adapters.apna import _clean_token
    from app.services.app_settings import set_setting

    token = _clean_token(payload.token)
    if not token:
        raise HTTPException(status_code=400, detail="Please paste your Apna token first.")

    await set_setting(db, "apna_token", token)
    if payload.org_id and payload.org_id.strip():
        await set_setting(db, "apna_org_id", payload.org_id.strip())
    await db.commit()
    logger.info("Apna token saved to server settings (org_id set=%s)", bool(payload.org_id))
    return {"saved": True}


@router.get("/apna-token")
async def get_apna_token_status(db: AsyncSession = Depends(get_db)):
    """Tell the page whether a server-side Apna token is saved — never the token itself."""
    from app.models.app_setting import AppSetting
    from app.services.app_settings import get_setting

    token = await get_setting(db, "apna_token")
    org_id = await get_setting(db, "apna_org_id")
    row = await db.get(AppSetting, "apna_token")
    return {
        "has_token": bool(token),
        "org_id": org_id,
        "updated_at": row.updated_at.isoformat() + "Z" if row and row.updated_at else None,
    }


@router.post("/apna-search")
async def apna_search(payload: ApnaSearchRequest):
    """Search Apna Hire's white-collar DB server-side (no CORS) and return preview list."""
    import httpx
    from app.adapters.apna import ApnaAdapter, _extract_live, _extract_total, to_preview

    adapter = ApnaAdapter(token=payload.apna_token, org_id=payload.org_id)
    try:
        data = await adapter.search_raw(
            keywords=payload.keywords, location=payload.location,
            experience_min=payload.experience_min, experience_max=payload.experience_max,
            page=payload.page, size=payload.size,
        )
    except httpx.HTTPStatusError as exc:
        raise HTTPException(status_code=exc.response.status_code,
                            detail=f"Apna API {exc.response.status_code}: {exc.response.text[:300]}")
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Could not reach Apna API: {str(exc)[:300]}")

    raw_list = _extract_live(data)
    total = _extract_total(data) or len(raw_list)
    return {"total": total, "page": payload.page, "size": payload.size,
            "candidates": [to_preview(c) for c in raw_list]}


@router.post("/apna-import", response_model=ImportResult)
async def apna_import(payload: ApnaImportRequest, db: AsyncSession = Depends(get_db)):
    """Import selected Apna candidates (raw dicts from /apna-search) into the pipeline."""
    from app.adapters.apna import _live_to_raw

    if not payload.candidates:
        raise HTTPException(status_code=400, detail="No candidates provided")

    raw = [rc for rc in (_live_to_raw(c) for c in payload.candidates) if rc]
    if not raw:
        raise HTTPException(status_code=422, detail="Could not parse any candidate from the provided data")

    logger.info("Apna live import: job=%d candidates=%d", payload.job_id, len(raw))
    return await _run_import_pipeline(raw, payload.job_id, payload.auto_outreach, db)

@router.post("/ai-screen/{job_id}")
async def ai_screen_pending(
    job_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Re-run AI screening on all PENDING (manual review) candidates for a job."""
    import asyncio
    from sqlalchemy import select
    from app.models.job import Job
    from app.models.candidate import Candidate
    from app.models.shortlist import ShortlistEntry, ShortlistStatus
    from app.services.ai_scoring import ai_score_candidate
    from app.services.scoring import score_candidate

    job_res = await db.execute(select(Job).where(Job.id == job_id))
    job = job_res.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    sl_res = await db.execute(
        select(ShortlistEntry).where(
            ShortlistEntry.job_id == job_id,
            ShortlistEntry.status.in_([ShortlistStatus.PENDING, ShortlistStatus.SHORTLISTED]),
        )
    )
    entries = sl_res.scalars().all()
    if not entries:
        return {"screened": 0, "message": "No pending candidates to screen"}

    cand_ids = [e.candidate_id for e in entries]
    c_res = await db.execute(select(Candidate).where(Candidate.id.in_(cand_ids)))
    candidates = {c.id: c for c in c_res.scalars().all()}

    entry_map = {e.candidate_id: e for e in entries}

    sem = asyncio.Semaphore(5)
    results = []

    async def _screen_one(cid: int):
        async with sem:
            candidate = candidates.get(cid)
            entry = entry_map.get(cid)
            if not candidate or not entry:
                return
            ai_result = await ai_score_candidate(candidate, job)
            if not ai_result or "score" not in ai_result:
                results.append({
                    "name": candidate.name,
                    "score": entry.score,
                    "result": "UNCHANGED",
                    "strengths": [],
                    "concerns": [],
                    "reasoning": "AI scoring not available",
                })
                return

            ai_score_100 = ai_result["score"] * 10
            old_score = entry.score or 0
            blended = round(old_score * 0.4 + ai_score_100 * 0.6, 2)
            decision = "AUTO_SHORTLIST" if blended >= 65 else ("MANUAL_REVIEW" if blended >= 40 else "REJECT")

            entry.score = blended
            existing_breakdown = entry.score_breakdown or {}
            entry.score_breakdown = {
                **existing_breakdown,
                "ai_strengths": ai_result.get("strengths", []),
                "ai_concerns": ai_result.get("concerns", []),
                "ai_reasoning": ai_result.get("reasoning", ""),
                "ai_opener": ai_result.get("personalized_opener", ""),
            }

            status_map = {
                "AUTO_SHORTLIST": ShortlistStatus.SHORTLISTED,
                "MANUAL_REVIEW": ShortlistStatus.PENDING,
                "REJECT": ShortlistStatus.REJECTED,
            }
            # Never silently demote a candidate who has already advanced past
            # review (CONTACTED/INTERESTED/INTERVIEW/HIRED) or who already has
            # outreach logged — re-screening must only (re)classify PENDING ones.
            if decision in status_map and entry.status in (
                ShortlistStatus.PENDING, ShortlistStatus.SHORTLISTED
            ):
                new_status = status_map[decision]
                # Don't downgrade a SHORTLISTED entry to REJECTED automatically.
                if not (entry.status == ShortlistStatus.SHORTLISTED
                        and new_status == ShortlistStatus.REJECTED):
                    entry.status = new_status

            results.append({
                "name": candidate.name,
                "candidate_id": candidate.id,
                "score": blended,
                "result": decision,
                "strengths": ai_result.get("strengths", []),
                "concerns": ai_result.get("concerns", []),
                "reasoning": ai_result.get("reasoning", ""),
                "opener": ai_result.get("personalized_opener", ""),
            })

    await asyncio.gather(*[_screen_one(cid) for cid in cand_ids])
    await db.commit()

    promoted = sum(1 for r in results if r.get("result") == "AUTO_SHORTLIST")
    rejected  = sum(1 for r in results if r.get("result") == "REJECT")

    return {
        "screened": len(results),
        "promoted_to_shortlist": promoted,
        "rejected": rejected,
        "stayed_pending": len(results) - promoted - rejected,
        "details": results,
  }
